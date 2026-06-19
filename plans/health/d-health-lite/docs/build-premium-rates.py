"""Regenerate dhl-premium-rates.js from premium rate dhl.xlsx"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(__file__).resolve().parent / "premium rate dhl.xlsx"
OUT_JS = ROOT / "dhl-premium-rates.js"

GENDER_MAP = {
    "เพศหญิง": "หญิง",
    "เพศชาย": "ชาย",
    "หญิง": "หญิง",
    "ชาย": "ชาย",
}


def load_rates() -> dict:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rates_by_gender: dict[str, dict[str, list[int]]] = {}
    ages: list[int] = []

    for sheet_name in wb.sheetnames:
        gender = GENDER_MAP.get(sheet_name, sheet_name)
        ws = wb[sheet_name]
        gender_rates: dict[str, list[int]] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            age = int(row[0])
            gender_rates[str(age)] = [int(row[i]) for i in range(1, 5)]
            ages.append(age)

        rates_by_gender[gender] = gender_rates

    if not ages:
        raise SystemExit("No premium rows found in workbook.")

    return {
        "minAge": min(ages),
        "maxAge": max(ages),
        "rates": rates_by_gender,
        "source": XLSX.name,
        "generatedAt": datetime.fromtimestamp(XLSX.stat().st_mtime).isoformat(),
    }


def write_js(payload: dict) -> None:
    body = json.dumps(payload["rates"], ensure_ascii=False, separators=(",", ":"))
    content = f"""/* D Health Lite premium rates — auto-generated from docs/{payload["source"]} */
/* Generated: {payload["generatedAt"]} */
window.DHL_PREMIUM_RATES = {{
  minAge: {payload["minAge"]},
  maxAge: {payload["maxAge"]},
  source: {json.dumps(payload["source"], ensure_ascii=False)},
  generatedAt: {json.dumps(payload["generatedAt"], ensure_ascii=False)},
  /** gender -> age -> [plan1m2k, plan5m4k, plan5m6k, plan5m8k] (THB/year) */
  rates: {body}
}};
"""
    OUT_JS.write_text(content, encoding="utf-8")


def main() -> None:
    payload = load_rates()
    write_js(payload)
    print(f"Wrote {OUT_JS} ({payload['minAge']}-{payload['maxAge']}, {len(payload['rates'])} genders)")


if __name__ == "__main__":
    main()
