"""Regenerate ehp-premium-rates.js from premium rate ehp.xlsx"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(__file__).resolve().parent / "premium rate ehp.xlsx"
OUT_JS = ROOT / "ehp-premium-rates.js"

PLANS = ["20m", "40m", "75m", "100m"]
AREAS = ["th", "asia", "world", "world-usa"]
AREA_SUFFIX = ["TH", "Asia", "WorldExUS", "World"]


def header_index(headers: list) -> dict[str, int]:
    return {str(name): idx for idx, name in enumerate(headers) if name}


def load_main_rates(ws) -> tuple[dict[str, dict], list[int]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = header_index(headers)
    rates: dict[str, dict] = {}
    ages: list[int] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        age = int(row[0])
        ages.append(age)
        plan_rates: dict[str, dict[str, int]] = {}

        for plan, plan_no in zip(PLANS, range(1, 5)):
            area_rates: dict[str, int] = {}
            for area, suffix in zip(AREAS, AREA_SUFFIX):
                key = f"Plan{plan_no}_{suffix}"
                idx = col[key]
                area_rates[area] = int(row[idx])
            plan_rates[plan] = area_rates

        rates[str(age)] = plan_rates

    return rates, ages


def load_addon_rates(ws, plan_keys: list[str]) -> tuple[dict[str, dict[str, int]], list[int]]:
    ages: list[int] = []
    rates: dict[str, dict[str, int]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        age = int(row[0])
        ages.append(age)
        rates[str(age)] = {
            plan_keys[0]: int(row[1]),
            plan_keys[1]: int(row[2]),
        }

    return rates, ages


def load_rates() -> dict:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    main_rates, main_ages = load_main_rates(wb["EHP"])
    maternity_rates, maternity_ages = load_addon_rates(
        wb["Maternity plus"], ["2m", "4m"]
    )
    wellbeing_rates, wellbeing_ages = load_addon_rates(
        wb["Wellbeing plus"], ["wb1", "wb2"]
    )

    if not main_ages:
        raise SystemExit("No premium rows found in EHP sheet.")

    all_ages = main_ages + maternity_ages + wellbeing_ages

    return {
        "minAge": min(main_ages),
        "maxAge": max(main_ages),
        "maternityMinAge": min(maternity_ages) if maternity_ages else 15,
        "maternityMaxAge": max(maternity_ages) if maternity_ages else 49,
        "rates": main_rates,
        "maternity": maternity_rates,
        "wellbeing": wellbeing_rates,
        "source": XLSX.name,
        "generatedAt": datetime.fromtimestamp(XLSX.stat().st_mtime).isoformat(),
    }


def write_js(payload: dict) -> None:
    rates_json = json.dumps(payload["rates"], ensure_ascii=False, separators=(",", ":"))
    maternity_json = json.dumps(
        payload["maternity"], ensure_ascii=False, separators=(",", ":")
    )
    wellbeing_json = json.dumps(
        payload["wellbeing"], ensure_ascii=False, separators=(",", ":")
    )
    content = f"""/* Elite Health Plus premium rates — auto-generated from docs/{payload["source"]} */
/* Generated: {payload["generatedAt"]} */
window.EHP_PREMIUM_RATES = {{
  hasRates: true,
  minAge: {payload["minAge"]},
  maxAge: {payload["maxAge"]},
  maternityMinAge: {payload["maternityMinAge"]},
  maternityMaxAge: {payload["maternityMaxAge"]},
  source: {json.dumps(payload["source"], ensure_ascii=False)},
  generatedAt: {json.dumps(payload["generatedAt"], ensure_ascii=False)},
  /** age -> plan -> area -> premium (THB/year) */
  rates: {rates_json},
  /** age -> 2m | 4m maternity add-on premium */
  maternity: {maternity_json},
  /** age -> wb1 | wb2 wellbeing add-on premium */
  wellbeing: {wellbeing_json}
}};
"""
    OUT_JS.write_text(content, encoding="utf-8")


def main() -> None:
    payload = load_rates()
    write_js(payload)
    print(
        f"Wrote {OUT_JS} (ages {payload['minAge']}-{payload['maxAge']}, "
        f"{len(payload['rates'])} main rows)"
    )


if __name__ == "__main__":
    main()
